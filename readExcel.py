import openpyxl

def czytaj_arkusz(plik, arkusz):
    wb = openpyxl.load_workbook(plik, data_only=True)
    sheet = wb.get_sheet_by_name(arkusz)
    data = []
    for row in sheet.iter_rows(min_row=2, max_row= sheet.max_row,min_col=2, max_col= sheet.max_column):
        data.append([cell.value for cell in row])
    return data

def czytajProfileStandardowe(file):
    sheets = {}
    profilB = czytaj_arkusz(file, "B11, B11em i B12")
    sheets["B11"] = profilB
    sheets["B11em"] = profilB
    sheets["B12"] = profilB
    profilC = czytaj_arkusz(file," C11, C11em i  C11s")
    sheets["C11"] = profilC
    sheets["C11em"] = profilC
    sheets["C11s"] = profilC
    sheets["C11o"] = czytaj_arkusz(file,"C11o")
    sheets["C12a"] = czytaj_arkusz(file,"C12a")
    sheets["C12b"] = czytaj_arkusz(file,"C12b")
    sheets["G11"] = czytaj_arkusz(file,"G11")
    sheets["G12"] = czytaj_arkusz(file,"G12")
    sheets["G12w"] = czytaj_arkusz(file,"G12w")
    sheets["G12as"] = czytaj_arkusz(file,"G12as")
    return sheets
